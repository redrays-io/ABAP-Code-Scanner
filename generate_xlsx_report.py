# generate_xlsx_report.py

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from typing import List, NamedTuple
import re

class ScanResult(NamedTuple):
    file_path: str
    line_number: int
    title: str
    message: str
    severity: str


def sanitize_for_excel(text):
    illegal_characters_pattern = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
    return illegal_characters_pattern.sub('', str(text))

def severity_key(result: ScanResult):
    severity_order = {
        "Critical": 1,
        "High": 2,
        "Medium": 3,
        "Low": 4,
        "Info": 5
    }
    return severity_order.get(result.severity, 6)

def generate_xlsx_report(results: List[ScanResult], output_file: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Security Scan Results"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    centered_alignment = Alignment(horizontal="center", vertical="center")
    wrapped_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    severity_colors = {
        "Critical": "FF0000",  # Red
        "High": "FFA500",  # Orange
        "Medium": "FFFF00",  # Yellow
        "Low": "90EE90",  # Light Green
        "Info": "ADD8E6"  # Light Blue
    }

    # Write headers
    headers = ["Severity", "Title", "File Path", "Line Number", "Message"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered_alignment

    # Write data
    for row, result in enumerate(results, start=2):
        ws.cell(row=row, column=1, value=sanitize_for_excel(result.severity)).alignment = wrapped_alignment
        ws.cell(row=row, column=2, value=result.title).alignment = wrapped_alignment
        ws.cell(row=row, column=3, value=sanitize_for_excel(result.file_path)).alignment = wrapped_alignment
        ws.cell(row=row, column=4, value=sanitize_for_excel(result.line_number)).alignment = wrapped_alignment
        ws.cell(row=row, column=5, value=sanitize_for_excel(result.message)).alignment = wrapped_alignment

        # Apply color to severity cell
        severity_cell = ws.cell(row=row, column=1)
        if result.severity in severity_colors:
            severity_cell.fill = PatternFill(start_color=severity_colors[result.severity],
                                             end_color=severity_colors[result.severity],
                                             fill_type="solid")

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

        # Set a maximum width for the message column
        message_column = get_column_letter(headers.index("Message") + 1)
        ws.column_dimensions[message_column].width = 50  # Adjust this value as needed

        # Set a maximum width for the check name column
        title = get_column_letter(headers.index("Title") + 1)
        ws.column_dimensions[title].width = 50  # Adjust this value as needed

        # Set a maximum width for the file path column
        file_path = get_column_letter(headers.index("File Path") + 1)
        ws.column_dimensions[file_path].width = 50  # Adjust this value as needed

    # Add filters
    ws.auto_filter.ref = ws.dimensions

    # Save the workbook
    wb.save(output_file)

if __name__ == "__main__":
    # Example usage
    sample_results = [
        ScanResult("file1.abap", 10, "Potential XSS", "Unsanitized input", "High"),
        ScanResult("file2.abap", 25, "SQL Injection", "Dynamic SQL query", "Critical"),
        # Add more sample results as needed
    ]
    generate_xlsx_report(sample_results, "sample_security_scan_report.xlsx")
    print("Sample report generated: sample_security_scan_report.xlsx")
